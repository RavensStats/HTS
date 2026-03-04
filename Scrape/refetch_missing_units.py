"""Re-fetch HTML files whose unit is missing from Extracted_Heroclix.xlsx.

A unit is considered "missing" when either:
  1. Its .txt file in unit_html_selenium/ exists but contains no rendered card
     data (i.e. lacks 'unitCardContainer' - captured before JS rendered), OR
  2. Its .txt file exists (with valid data) but there is no corresponding row
     in the Units sheet of Extracted_Heroclix.xlsx.

The script loads the xlsx, builds a set of known unit IDs (set_abbr + set_number),
then inspects every file in unit_html_selenium/ and re-fetches any that are missing
from the xlsx or that have empty/invalid content.

Usage:
    python refetch_missing_units.py
"""

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import InvalidSessionIdException, WebDriverException
import time
import os
import re
import openpyxl


# ---------------------------------------------------------------------------
# Config (must match save_unit_html_selenium.py)
# ---------------------------------------------------------------------------
BASE_URL = "https://hcunits.net/units/"
OUTPUT_DIR = "unit_html_selenium"
EXTRACTED_XLSX = "Extracted_Heroclix.xlsx"

JS_RENDER_TIMEOUT = 5
COURTESY_DELAY = 0.5


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def file_has_unit_data(path: str) -> bool:
    """Return True if the saved HTML contains a rendered unitCardContainer."""
    try:
        with open(path, encoding="utf-8") as f:
            content = f.read()
        return "unitCardContainer" in content
    except Exception:
        return False


def load_extracted_unit_ids(xlsx_path: str) -> set:
    """Read the Units sheet and return a set of '<set_abbr><set_number>' strings."""
    unit_ids: set = set()
    if not os.path.exists(xlsx_path):
        print(f"  Warning: {xlsx_path} not found - treating all files as missing.")
        return unit_ids
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Units" not in wb.sheetnames:
        print(f"  Warning: 'Units' sheet not found in {xlsx_path}.")
        wb.close()
        return unit_ids
    ws = wb["Units"]
    # Row 1 is the header; data starts at row 2.
    # Column A = Set Abbreviation, Column C = Set Number.
    for row in ws.iter_rows(min_row=2, values_only=True):
        set_abbr = str(row[0]).strip() if row[0] is not None else ""
        set_number = str(row[2]).strip() if row[2] is not None else ""
        if set_abbr and set_number:
            unit_ids.add(f"{set_abbr}{set_number}")
    wb.close()
    print(f"  Loaded {len(unit_ids)} extracted unit IDs from {xlsx_path}.")
    return unit_ids


def load_page_with_unit_card(driver, unit_code: str, timeout: int = JS_RENDER_TIMEOUT):
    """Navigate to a unit page, wait for the card JS to render. Returns page source or None."""
    url = f"{BASE_URL}{unit_code}/"
    driver.get(url)
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "unitCardsContainer"))
        )
        try:
            WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.CLASS_NAME, "unitCardContainer"))
            )
        except Exception:
            pass
        time.sleep(COURTESY_DELAY)
        return driver.page_source
    except Exception:
        time.sleep(COURTESY_DELAY)
        return None


def save_unit(unit_code: str, html: str):
    out_path = os.path.join(OUTPUT_DIR, f"{unit_code}.txt")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


# ---------------------------------------------------------------------------
# Build the list of unit IDs to re-fetch
# ---------------------------------------------------------------------------
print("=" * 60)
print("Refetch Missing Units")
print("=" * 60)

os.makedirs(OUTPUT_DIR, exist_ok=True)
extracted_ids = load_extracted_unit_ids(EXTRACTED_XLSX)

txt_files = sorted(f for f in os.listdir(OUTPUT_DIR) if f.endswith(".txt"))
print(f"  Found {len(txt_files)} file(s) in '{OUTPUT_DIR}'.")

to_refetch: list[str] = []  # unit codes (= file base names) to re-fetch

for filename in txt_files:
    base = os.path.splitext(filename)[0]  # e.g. 'dicn001'
    path = os.path.join(OUTPUT_DIR, filename)

    missing_from_xlsx = base not in extracted_ids
    has_bad_content = not file_has_unit_data(path)

    if has_bad_content:
        # File was captured before JS rendered - always needs a fresh fetch.
        reason = "empty content"
        if missing_from_xlsx:
            reason += " + not in xlsx"
        print(f"  Will refetch [{reason}]: {base}")
        to_refetch.append(base)
    elif missing_from_xlsx:
        # File has valid content but no xlsx row yet - needs re-extraction only,
        # no need to re-fetch from the web.
        print(f"  Skipping [good content, not in xlsx - re-run extraction]: {base}")

if not to_refetch:
    print("\nNothing to re-fetch - all files are present in the xlsx.")
    exit(0)

print(f"\n{len(to_refetch)} unit(s) to re-fetch.")

# ---------------------------------------------------------------------------
# Launch Selenium and re-fetch
# ---------------------------------------------------------------------------
def make_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--headless")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1920,1080")
    return webdriver.Chrome(options=opts)


driver = make_driver()

BAR_WIDTH = 40

def print_progress(current: int, total: int, unit_code: str, suffix: str = "") -> None:
    """Print an overwriting progress bar to stdout."""
    pct = current / total
    filled = int(BAR_WIDTH * pct)
    bar = "#" * filled + "-" * (BAR_WIDTH - filled)
    label = f"{unit_code:<20}" + (f" {suffix}" if suffix else "")
    print(f"\r  [{bar}] {pct:5.1%} ({current}/{total})  {label}", end="", flush=True)


failed: list[str] = []
total = len(to_refetch)
for idx, unit_code in enumerate(to_refetch, start=1):
    print_progress(idx, total, unit_code)
    # Retry once after a session/driver crash (e.g. caused by power-save sleep).
    for attempt in range(2):
        try:
            html = load_page_with_unit_card(driver, unit_code)
            break
        except (InvalidSessionIdException, WebDriverException) as exc:
            if attempt == 0:
                print(f"\n    Driver session lost ({type(exc).__name__}), restarting driver...")
                try:
                    driver.quit()
                except Exception:
                    pass
                driver = make_driver()
            else:
                print(f"\n    Driver restart failed for {unit_code}: {exc}")
                html = None
                break
    if html and "unitCardContainer" in html:
        print_progress(idx, total, unit_code, suffix="OK")
        save_unit(unit_code, html)
    else:
        print_progress(idx, total, unit_code, suffix="NOT FOUND")
        print(f"\n    Unit card not found for {unit_code}, skipping.")
        failed.append(unit_code)

try:
    driver.quit()
except Exception:
    pass

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
print("\n\n" + "=" * 60)
print(f"Re-fetched : {len(to_refetch) - len(failed)}")
print(f"Still failed: {len(failed)}")
if failed:
    print("Failed units:")
    for u in failed:
        print(f"  {u}")
print("Done.")
