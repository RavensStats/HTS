import requests
import os
import re
from bs4 import BeautifulSoup
import openpyxl


def _split_set_abbr_number(base):
    """Split a file's base name into (set_abbr, set_number).

    The unit number is the trailing 3-digit block at the end of the name
    (matched by ``\\d{3}[a-z0-9.]*$``), optionally preceded by one or more
    consecutive uppercase letters that act as a sub-type prefix (M, B, H, S,
    F, G, L, TM, TA, …).  The set abbreviation is everything before that,
    with any trailing dash stripped.

    Falls back to the leading letter-only prefix for very short numeric
    suffixes (e.g. 'anW-1').

    Examples
    --------
    'trek3001'    -> ('trek3',   '001')
    'roc18M001'   -> ('roc18',   'M001')
    'smbaTM001'   -> ('smba',    'TM001')
    'gotg2m001'   -> ('gotg2m',  '001')   # lowercase stays in abbr
    'gc001'       -> ('gc',      '001')
    'av4e001.1'   -> ('av4e',    '001.1')
    'micnMVX-010' -> ('micnMVX', '010')
    'anW-1'       -> ('anW',     '-1')    # fallback path
    """
    # Strip optional dotted revision suffix (.1, .01, .1s ...) before searching.
    versioned = re.sub(r'\.\d+[a-z]*$', '', base)
    # Find the leftmost position where the end-anchored unit-number pattern fits.
    # [A-Z]*  = optional uppercase sub-type prefix (M, B, TM, TA, ...)
    # \d{3}   = exactly 3 digits (never greedily swallows a set-suffix digit)
    # [a-z]*  = optional lowercase variant suffix (a, b, bt, bta, ...)
    m = re.search(r'[A-Z]*\d{3}[a-z]*$', versioned)
    if m:
        pos = m.start()                      # includes any uppercase prefix chars
        set_abbr = base[:pos].rstrip('-')
        set_number = base[pos:]              # use original to preserve .N suffix
        return set_abbr, set_number
    # No 3-digit tail - use leading letter-only prefix (handles e.g. 'anW-1').
    m2 = re.match(r'^([a-zA-Z]+)', base)
    set_abbr = m2.group(1) if m2 else base
    return set_abbr, base[len(set_abbr):]


INPUT_DIR = "unit_html_selenium"
OUTPUT_FILE = "Extracted_Heroclix.xlsx"
MAX_CLICKS = 30  # Maximum dial clicks (columns) to capture

# Known improved ability SVG stems, in display order
IMPROVED_MOVEMENT_ABILITIES = [
    "elevated", "hindering", "blocking", "destroy_blocking",
    "characters", "move_through", "water", "indoor", "outdoor",
]
IMPROVED_TARGETING_ABILITIES = [
    "elevated", "hindering", "blocking", "characters", "adjacent",
]


# ---------------------------------------------------------------------------
# Helper: extract special powers (up to 10)
# ---------------------------------------------------------------------------
def extract_special_powers(unit_cards_container):
    special_names = []
    special_abilities = []
    special_types = []
    # Normal hero/character cards use largeCardSpecialPowersTableNormalDial.
    # F-type cards (Feat / Object / NoDial) use largeCardSpecialPowersTableNoDial.
    main_powers_section = unit_cards_container.find("div", id="largeCardSpecialPowersTableNormalDial")
    if not main_powers_section:
        main_powers_section = unit_cards_container.find("div", id="largeCardSpecialPowersTableNoDial")
    if not main_powers_section:
        return ["" for _ in range(10)], ["" for _ in range(10)], ["" for _ in range(10)]
    titles = main_powers_section.find_all("div", class_="largeCardSpecialPowerTitle")
    for title_div in titles:
        icon_div = title_div.find("div", class_="specialPowerIcon")
        if icon_div and icon_div.find("img", src="https://storage.googleapis.com/static.hcunits.net/images/sp/trait.svg"):
            continue
        name_div = title_div.find("div", class_="largeCardSpecialPowerName")
        name = name_div.get_text(strip=True) if name_div else ""
        desc_div = title_div.find_next_sibling("div", class_="largeCardSpecialPowerDescription")
        ability = ""
        if desc_div:
            parts = []
            for elem in desc_div.descendants:
                if isinstance(elem, str):
                    parts.append(elem)
                elif elem.name == "img" and elem.get("src"):
                    parts.append(f"{{IMG:{elem['src']}}}")
            ability = "".join(parts)
        type_val = ""
        if icon_div:
            icon_img = icon_div.find("img", class_="specialPowerIconCombatSymbol")
            if icon_img:
                src = icon_img.get("src", "")
                type_val = os.path.splitext(os.path.basename(src))[0]
        special_names.append(name)
        special_abilities.append(ability)
        special_types.append(type_val)
        if len(special_names) >= 10:
            break
    while len(special_names) < 10:
        special_names.append("")
        special_abilities.append("")
        special_types.append("")
    return special_names, special_abilities, special_types


# ---------------------------------------------------------------------------
# Helper: extract traits (up to 10)
# ---------------------------------------------------------------------------
def extract_traits(unit_cards_container):
    traits = []
    trait_abilities = []
    trait_imgs = unit_cards_container.find_all("img", src="https://storage.googleapis.com/static.hcunits.net/images/sp/trait.svg")
    for img in trait_imgs[:10]:
        icon_div = img.find_parent("div", class_="specialPowerIcon")
        title_div = icon_div.find_parent("div", class_="largeCardSpecialPowerTitle") if icon_div else None
        trait_name = ""
        trait_ability = ""
        if title_div:
            name_div = title_div.find("div", class_="largeCardSpecialPowerName")
            if name_div:
                trait_name = name_div.get_text(strip=True)
            desc_div = title_div.find_next_sibling("div", class_="largeCardSpecialPowerDescription")
            if desc_div:
                parts = []
                for elem in desc_div.descendants:
                    if isinstance(elem, str):
                        parts.append(elem)
                    elif elem.name == "img" and elem.get("src"):
                        parts.append(f"{{IMG:{elem['src']}}}")
                trait_ability = "".join(parts)
        traits.append(trait_name)
        trait_abilities.append(trait_ability)
    while len(traits) < 10:
        traits.append("")
        trait_abilities.append("")
    return traits, trait_abilities


# ---------------------------------------------------------------------------
# Helper: extract improved movement / targeting abilities as individual booleans
# ---------------------------------------------------------------------------
def extract_improved_abilities(unit_cards_container):
    """Returns an ordered dict {field_name: 'True'/'False'} for all improved abilities."""
    result = {}
    for stem in IMPROVED_MOVEMENT_ABILITIES:
        result[f"Improved Movement {stem.replace('_', ' ').title()}"] = "False"
    for stem in IMPROVED_TARGETING_ABILITIES:
        result[f"Improved Targeting {stem.replace('_', ' ').title()}"] = "False"

    improved_div = unit_cards_container.find("div", id="largeCardImprovedAbilities")
    if not improved_div:
        return result

    current_category = None
    for child in improved_div.children:
        if not hasattr(child, "name") or not child.name:
            continue
        if child.name == "img" and "largeCardImprovedAbility" in child.get("class", []):
            src = child.get("src", "")
            if "movement.svg" in src:
                current_category = "Movement"
            elif "targeting.svg" in src:
                current_category = "Targeting"
        elif child.name == "div" and "largeCardImprovedAbility" in child.get("class", []):
            if current_category:
                img = child.find("img", class_="largeCardImprovedAbilityImg")
                if img:
                    stem = img.get("src", "").rsplit("/", 1)[-1].replace(".svg", "")
                    field = f"Improved {current_category} {stem.replace('_', ' ').title()}"
                    if field in result:
                        result[field] = "True"
                    else:
                        # Unknown ability — add it dynamically
                        result[field] = "True"
                        print(f"  Note: unknown improved ability field: {field}")
    return result


# ---------------------------------------------------------------------------
# Helper: combat symbol type from img element
# ---------------------------------------------------------------------------
def cs_type(img):
    src = img.get("src", "")
    stem = src.rsplit("/", 1)[-1].replace(".svg", "")
    return stem.replace("-", " ").title() if stem else ""


# ---------------------------------------------------------------------------
# Helper: dial row -> values + powers (padded to MAX_CLICKS each)
# Accepts either a single row element or a list of <td> elements directly.
# ---------------------------------------------------------------------------
def extract_dial_row(row_or_tds, num_clicks):
    values, powers = [], []
    tds = row_or_tds if isinstance(row_or_tds, list) else row_or_tds.find_all("td")
    for td in tds[:num_clicks]:
        values.append(td.get_text(strip=True))
        power = ""
        title = td.get("data-mdb-original-title", "")
        if title:
            s2 = BeautifulSoup(title, "html.parser")
            bold = s2.find("b")
            if bold:
                power = bold.get_text(strip=True)
        powers.append(power)
    while len(values) < MAX_CLICKS:
        values.append("")
        powers.append("")
    return values + powers


# ---------------------------------------------------------------------------
# DOM path / aria-label helpers (used inside add_element_and_values)
# ---------------------------------------------------------------------------
def get_dom_path(element):
    path = []
    while element and element.name:
        sibs = list(element.parent.find_all(element.name, recursive=False)) if element.parent else []
        idx = sibs.index(element) if sibs else 0
        path.append(f"{element.name}[{idx}]")
        element = element.parent
    return "/".join(reversed(path))


def parse_aria_label(val):
    if val and "<b>" in val and "</b>" in val:
        soup2 = BeautifulSoup(val, "html.parser")
        bold = soup2.find("b")
        team_ability = bold.text.strip() if bold else ""
        description = soup2.get_text().replace(team_ability, "", 1).lstrip(": ").strip()
        return team_ability, description
    return "", ""


# ---------------------------------------------------------------------------
# Recursive element walker
# ---------------------------------------------------------------------------
def add_element_and_values(el):
    path = get_dom_path(el)
    text = el.get_text(strip=True)
    if el.name == "div" and el.get("id", "") == "largeCardKeywords":
        keywords = ", ".join([a.get_text(strip=True) for a in el.find_all("a")])
        _scratch.append([path, el.name, "text", keywords, "Keywords", keywords])
        add_element_and_values.summary_fields["Keywords"] = keywords
    elif el.name == "div" and el.get("id", "") == "largeCardRealName":
        real_name = text.replace("REAL NAME:", "").strip()
        _scratch.append([path, el.name, "text", text, "Real Name", real_name])
        add_element_and_values.summary_fields["Real Name"] = real_name
    elif el.name == "img" and "largeCardSetIcon" in el.get("class", []):
        set_name = el.get("aria-label", "").strip() or el.get("data-mdb-original-title", "").strip()
        if set_name:
            _scratch.append([path, el.name, "aria-label", set_name, "Set Name", set_name])
            add_element_and_values.summary_fields["Set Name"] = set_name
    elif el.name == "div" and "largeCardPointValue" in el.get("class", []) and text.isdigit():
        _scratch.append([path, el.name, "text", text, "Points", text])
        add_element_and_values.summary_fields["Points"] = text
    elif el.name == "div" and "largeCardRange" in el.get("class", []):
        range_val = "".join(filter(str.isdigit, text))
        _scratch.append([path, el.name, "text", text, "Range Value", range_val])
        add_element_and_values.summary_fields["Range Value"] = range_val
    elif text:
        _scratch.append([path, el.name, "text", text, "", ""])
    if el.name == "img" and "/imp/" in el.get("src", ""):
        src = el.get("src", "")
        match = re.search(r"/imp/([a-zA-Z0-9_\-]+)\.svg", src)
        if match:
            imp_type = match.group(1).capitalize()
            add_element_and_values.improved_movement_types.add(imp_type)
            if src == "https://storage.googleapis.com/static.hcunits.net/images/imp/elevated.svg":
                add_element_and_values.improved_movement_elevated_found = True
            add_element_and_values.summary_fields["Improved Movement Types"] = ", ".join(sorted(add_element_and_values.improved_movement_types))
    if el.name == "img" and el.get("src", "") == "https://storage.googleapis.com/static.hcunits.net/images/cs/bolt.svg":
        add_element_and_values.bolt_svg_count += 1
    for attr, val in el.attrs.items():
        if isinstance(val, list):
            val = ", ".join(val)
        if attr == "aria-label" and not (el.name == "img" and "largeCardSetIcon" in el.get("class", [])):
            team_ability, description = parse_aria_label(val)
            _scratch.append([path, el.name, attr, val, team_ability, description])
        elif attr != "aria-label":
            _scratch.append([path, el.name, attr, val, "", ""])
    for child in el.find_all(recursive=False):
        add_element_and_values(child)


add_element_and_values.improved_movement_elevated_found = False
add_element_and_values.improved_movement_types = set()
add_element_and_values.bolt_svg_count = 0
add_element_and_values.summary_fields = {}


# ---------------------------------------------------------------------------
# Workbook setup
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()

# Scratch sheet used internally by add_element_and_values, removed before save
_scratch = wb.active
_scratch.title = "_scratch"
_scratch.append(["element_path", "element_name", "attribute", "value", "Team Ability", "Description"])

# Units sheet - one row per unit
UNIT_HEADERS = [
    "Set Abbreviation", "Set Name", "Set Number",
    "Unit Name", "Base Size",
    "Starting Line 1", "Starting Line 2", "Starting Line 3", "Starting Line 4",
    "Starting Line 5", "Starting Line 6", "Starting Line 7", "Starting Line 8",
    "Starting Line 9", "Starting Line 10",
    "Point Value 1", "Point Value 2", "Point Value 3", "Point Value 4",
    "Point Value 5", "Point Value 6", "Point Value 7", "Point Value 8",
    "Point Value 9", "Point Value 10",
    "Trait 1", "Trait 1 Ability",
    "Trait 2", "Trait 2 Ability",
    "Trait 3", "Trait 3 Ability",
    "Trait 4", "Trait 4 Ability",
    "Trait 5", "Trait 5 Ability",
    "Trait 6", "Trait 6 Ability",
    "Trait 7", "Trait 7 Ability",
    "Trait 8", "Trait 8 Ability",
    "Trait 9", "Trait 9 Ability",
    "Trait 10", "Trait 10 Ability",
]
for i in range(1, 11):
    UNIT_HEADERS += [f"Special Power {i}", f"Special Power Ability {i}", f"Special Power Type {i}"]
UNIT_HEADERS += [
    "Keywords", "Real Name", "Range Value",
    "Movement Type", "Attack Type", "Defense Type", "Damage Type",
]
for _stem in IMPROVED_MOVEMENT_ABILITIES:
    UNIT_HEADERS.append(f"Improved Movement {_stem.replace('_', ' ').title()}")
for _stem in IMPROVED_TARGETING_ABILITIES:
    UNIT_HEADERS.append(f"Improved Targeting {_stem.replace('_', ' ').title()}")
UNIT_HEADERS.append("Number of Targets")
ws_units = wb.create_sheet(title="Units")
ws_units.append(UNIT_HEADERS)

# Dial sheet - one row per unit, MAX_CLICKS columns per stat
DIAL_HEADERS = ["Set Abbreviation", "Set Name", "Set Number"]
for stat in ["Speed", "Attack", "Defense", "Damage"]:
    for i in range(1, MAX_CLICKS + 1):
        DIAL_HEADERS.append(f"{stat} Value {i}")
    for i in range(1, MAX_CLICKS + 1):
        DIAL_HEADERS.append(f"{stat} Power {i}")
ws_dial = wb.create_sheet(title="Dial")
ws_dial.append(DIAL_HEADERS)


# ---------------------------------------------------------------------------
# Main loop - process every .txt file in INPUT_DIR
# ---------------------------------------------------------------------------
txt_files = sorted(f for f in os.listdir(INPUT_DIR) if f.endswith(".txt"))
print(f"Found {len(txt_files)} file(s) to process.")

for txt_filename in txt_files:
    input_path = os.path.join(INPUT_DIR, txt_filename)
    print(f"Processing {txt_filename}...")

    with open(input_path, encoding="utf-8") as f:
        html = f.read()
    soup = BeautifulSoup(html, "html.parser")

    # Derive set_abbr and set_number from filename (e.g. "trek3001" -> "trek3", "001")
    base = os.path.splitext(txt_filename)[0]
    set_abbr, set_number = _split_set_abbr_number(base)

    unit_cards_container = soup.find("div", id="unitCardsContainer")
    if not unit_cards_container:
        print(f"  unitCardsContainer not found, skipping.")
        continue

    # Guard against files saved before JavaScript rendered the card data.
    # In those files the #unitCards div is present but completely empty.
    if not unit_cards_container.find("div", class_="unitCardContainer"):
        print(f"  No unitCardContainer found - page was captured before JS rendered. Skipping (re-run save_unit_html_selenium.py to re-fetch).")
        continue

    # Reset function state for this unit
    add_element_and_values.improved_movement_elevated_found = False
    add_element_and_values.improved_movement_types = set()
    add_element_and_values.bolt_svg_count = 0
    add_element_and_values.summary_fields = {}

    # Direct field extraction
    trait_names, trait_abilities_list = extract_traits(unit_cards_container)
    special_power_names, special_power_abilities, special_power_types = extract_special_powers(unit_cards_container)

    kw_el = unit_cards_container.find("div", id="largeCardKeywords")
    keywords_val = ", ".join([a.get_text(strip=True) for a in kw_el.find_all("a")]) if kw_el else ""

    rn_el = unit_cards_container.find("div", id="largeCardRealName")
    real_name_val = rn_el.get_text(strip=True).replace("REAL NAME:", "").strip() if rn_el else ""

    si_el = unit_cards_container.find("img", class_="largeCardSetIcon")
    set_name_val = ""
    if si_el:
        set_name_val = si_el.get("aria-label", "").strip() or si_el.get("data-mdb-original-title", "").strip()

    rng_el = unit_cards_container.find("div", class_="largeCardRange")
    range_val = rng_el.get_text(strip=True) if rng_el else ""

    name_el = unit_cards_container.find("div", id="largeCardName")
    unit_name_val = name_el.get_text(strip=True) if name_el else ""

    dim_el = unit_cards_container.find("div", id="cardDimensions")
    oval_el = dim_el.find("div", class_="cardDimensionsOval") if dim_el else None
    base_size_val = oval_el.get_text(strip=True) if oval_el else "1x1"

    cs_imgs = unit_cards_container.find_all("img", class_="largeCardCombatSymbolImg")
    movement_type_val = cs_type(cs_imgs[0]) if len(cs_imgs) > 0 else ""
    attack_type_val   = cs_type(cs_imgs[1]) if len(cs_imgs) > 1 else ""
    defense_type_val  = cs_type(cs_imgs[2]) if len(cs_imgs) > 2 else ""
    damage_type_val   = cs_type(cs_imgs[3]) if len(cs_imgs) > 3 else ""

    # Run recursive walker (bolt count only; improved movement handled below)
    add_element_and_values(unit_cards_container)
    num_targets = str(add_element_and_values.bolt_svg_count)

    # Improved movement / targeting individual booleans
    improved_abilities = extract_improved_abilities(unit_cards_container)

    # Point values (one per starting line)
    pv_container = unit_cards_container.find("div", class_="largeCardPointValues")
    point_values = []
    if pv_container:
        for pv_div in pv_container.find_all("div", class_="largeCardPointValue"):
            point_values.append(pv_div.get_text(strip=True))
    else:
        # F-type cards (Feat / Object) store a single point value as plain text
        # inside a div with class largeCardObjectPointValues, e.g. "10 POINTS".
        obj_pv = unit_cards_container.find("div", class_="largeCardObjectPointValues")
        if obj_pv:
            pv_text = obj_pv.get_text(strip=True)
            # Strip the " POINTS" suffix to leave the raw number.
            pv_num = pv_text.replace("POINTS", "").strip()
            if pv_num:
                point_values.append(pv_num)
    while len(point_values) < 10:
        point_values.append("")

    # Dial extraction
    # Units with long dials split clicks across multiple largeCardDialTable elements
    # (e.g. 12 + 3 = 15 clicks).  Collect all TDs per stat row across every table.
    starting_line_clicks = []
    dial_row_data = []
    dial_tables = unit_cards_container.find_all("table", class_="largeCardDialTable")
    if dial_tables:
        # Merge the stat rows (indices 1-4) across all tables
        stat_tds = [[] for _ in range(4)]  # [speed_tds, attack_tds, defense_tds, damage_tds]
        for dial_table in dial_tables:
            rows = dial_table.find_all("tr", class_="largeCardDialRow")
            if len(rows) >= 5:
                for stat_idx in range(4):
                    stat_tds[stat_idx].extend(rows[stat_idx + 1].find_all("td"))
        num_clicks = min(len(stat_tds[0]), MAX_CLICKS)
        dial_table = dial_tables[0]  # used below for style/starting-line calcs
        for stat_idx in range(4):
            dial_row_data.extend(extract_dial_row(stat_tds[stat_idx], num_clicks))
        dial_table_width = 280
        mw = re.search(r"width\s*:\s*(\d+)px", dial_table.get("style", ""))
        if mw:
            dial_table_width = int(mw.group(1))
        # Starting line markers are positioned at fixed 23px steps starting at left=31px,
        # matching a 12-column reference grid (280px / 12 ≈ 23px), regardless of actual
        # number of clicks. Formula: click = round((left_px - 31) / 23) + 1
        DIAL_OFFSET_PX = 31
        DIAL_STEP_PX = 23
        for sl in unit_cards_container.find_all("div", class_="largeCardDialStartingLine")[:10]:
            ml = re.search(r"left\s*:\s*(\d+)px", sl.get("style", ""))
            if ml:
                left_px = int(ml.group(1))
                click = round((left_px - DIAL_OFFSET_PX) / DIAL_STEP_PX) + 1
                starting_line_clicks.append(min(max(click, 1), num_clicks))
            else:
                starting_line_clicks.append("")
    while len(starting_line_clicks) < 10:
        starting_line_clicks.append("")
    expected_dial_len = 4 * 2 * MAX_CLICKS
    while len(dial_row_data) < expected_dial_len:
        dial_row_data.append("")

    # Download set icon (once per unique set abbreviation)
    if si_el and set_abbr:
        set_icon_src = si_el.get("src", "").strip()
        if set_icon_src:
            set_images_dir = os.path.normpath(
                os.path.join(os.path.dirname(os.path.abspath(input_path)), "..", "Set Images")
            )
            os.makedirs(set_images_dir, exist_ok=True)
            dest_path = os.path.join(set_images_dir, f"{set_abbr}.svg")
            if not os.path.exists(dest_path):
                try:
                    resp = requests.get(set_icon_src, timeout=10)
                    resp.raise_for_status()
                    with open(dest_path, "wb") as fout:
                        fout.write(resp.content)
                    print(f"  Saved set icon -> {dest_path}")
                except Exception as e:
                    print(f"  Failed to download set icon: {e}")

    # Download unit card image, saved as <set_abbr><set_number>.<ext>
    unit_img_el = unit_cards_container.find("img", id="largeCardToken")
    if unit_img_el:
        unit_img_src = unit_img_el.get("src", "").strip()
        ext = os.path.splitext(unit_img_src)[-1] or ".png"
        unit_images_dir = os.path.normpath(
            os.path.join(os.path.dirname(os.path.abspath(input_path)), "..", "Unit_Images")
        )
        os.makedirs(unit_images_dir, exist_ok=True)
        unit_img_dest = os.path.join(unit_images_dir, f"{set_abbr}{set_number}{ext}")
        if not os.path.exists(unit_img_dest):
            try:
                resp = requests.get(unit_img_src, timeout=10)
                resp.raise_for_status()
                with open(unit_img_dest, "wb") as fout:
                    fout.write(resp.content)
                print(f"  Saved unit image -> {unit_img_dest}")
            except Exception as e:
                print(f"  Failed to download unit image: {e}")

    # Append Units row
    unit_row = [
        set_abbr, set_name_val, set_number,
        unit_name_val, base_size_val,
        starting_line_clicks[0], starting_line_clicks[1],
        starting_line_clicks[2], starting_line_clicks[3],
        starting_line_clicks[4], starting_line_clicks[5],
        starting_line_clicks[6], starting_line_clicks[7],
        starting_line_clicks[8], starting_line_clicks[9],
        point_values[0], point_values[1], point_values[2], point_values[3],
        point_values[4], point_values[5], point_values[6], point_values[7],
        point_values[8], point_values[9],
    ]
    for i in range(10):
        unit_row += [trait_names[i], trait_abilities_list[i]]
    for i in range(10):
        unit_row += [special_power_names[i], special_power_abilities[i], special_power_types[i]]
    unit_row += [
        keywords_val, real_name_val, range_val,
        movement_type_val, attack_type_val, defense_type_val, damage_type_val,
    ]
    for _stem in IMPROVED_MOVEMENT_ABILITIES:
        unit_row.append(improved_abilities.get(f"Improved Movement {_stem.replace('_', ' ').title()}", "False"))
    for _stem in IMPROVED_TARGETING_ABILITIES:
        unit_row.append(improved_abilities.get(f"Improved Targeting {_stem.replace('_', ' ').title()}", "False"))
    unit_row.append(num_targets)
    ws_units.append(unit_row)

    # Append Dial row
    ws_dial.append([set_abbr, set_name_val, set_number] + dial_row_data)

    print(f"  Done: {set_abbr} {set_number}  ({set_name_val})")

# Remove scratch sheet and save
wb.remove(_scratch)
wb.save(OUTPUT_FILE)
print(f"\nSaved all data to {OUTPUT_FILE}")
