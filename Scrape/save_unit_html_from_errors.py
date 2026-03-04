from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import os
import openpyxl


BASE_URL = "https://hcunits.net/units/"
OUTPUT_DIR = "unit_html_selenium"
ERROR_FILE = "save_unit_errors.xlsx"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Max seconds to wait for the unit card JS to render.
JS_RENDER_TIMEOUT = 5
# Courtesy pause between requests (seconds).
COURTESY_DELAY = 0.5


def load_page_with_unit_card(unit_code, timeout=None):
    """Navigate to a unit page and wait for unitCardsContainer. Returns page source or None."""
    if timeout is None:
        timeout = JS_RENDER_TIMEOUT
    url = f"{BASE_URL}{unit_code}/"
    driver.get(url)
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "unitCardsContainer"))
        )
        time.sleep(COURTESY_DELAY)
        return driver.page_source
    except Exception:
        time.sleep(COURTESY_DELAY)
        return None


def get_unit_ids_from_sidebar(html):
    """Parse the set sidebar and return all unit IDs listed."""
    soup = BeautifulSoup(html, "html.parser")
    unit_ids = []
    for tag in soup.find_all("a", id=lambda x: x and x.startswith("unitListPanelItem_")):
        unit_id = tag["id"].replace("unitListPanelItem_", "")
        unit_ids.append(unit_id)
    return unit_ids


def save_unit(unit_code, html):
    out_path = os.path.join(OUTPUT_DIR, f"{unit_code}.txt")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Saved {unit_code}")


# Read the error file - column A = set abbreviation, column D = manual entry point
wb = openpyxl.load_workbook(ERROR_FILE)
ws = wb.active

entries = []
for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
    set_abbr = row[0]  # Column A
    entry_point = row[3]  # Column D
    if set_abbr and entry_point:
        entries.append((str(set_abbr).strip(), str(entry_point).strip()))

if not entries:
    print("No entries with manual entry points found in column D of the error file.")
    exit()

print(f"Found {len(entries)} set(s) to process from error file.")

options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument('--no-sandbox')
options.add_argument('--window-size=1920,1080')

driver = webdriver.Chrome(options=options)

# Track any remaining failures
remaining_errors = []

for set_abbr, entry_point in entries:
    print(f"\nProcessing set: {set_abbr} (entry point: {entry_point})")

    # Step 1: Load the manually provided entry point.
    entry_html = load_page_with_unit_card(entry_point)
    if not entry_html:
        print(f"  Entry point {entry_point} failed to load, skipping.")
        remaining_errors.append((set_abbr, entry_point, "Entry point failed to load"))
        continue

    # Step 2: Extract the full unit ID list from the sidebar.
    unit_ids = get_unit_ids_from_sidebar(entry_html)
    if not unit_ids:
        print(f"  No unit IDs found in sidebar for {entry_point}, skipping.")
        remaining_errors.append((set_abbr, entry_point, "No sidebar IDs found"))
        continue
    print(f"  Found {len(unit_ids)} unit(s) in sidebar.")

    # Step 3: Fetch and save each unit, skipping ones already downloaded.
    for unit_id in unit_ids:
        out_path = os.path.join(OUTPUT_DIR, f"{unit_id}.txt")
        if os.path.exists(out_path):
            print(f"  Already exists, skipping: {unit_id}")
            continue
        html = load_page_with_unit_card(unit_id)
        if html:
            save_unit(unit_id, html)
        else:
            print(f"  Unit card not found for {unit_id}, skipping.")
            remaining_errors.append((set_abbr, unit_id, "Unit not found"))

driver.quit()

# Write remaining failures back to the error file on a new sheet
if remaining_errors:
    if "Remaining Errors" in wb.sheetnames:
        del wb["Remaining Errors"]
    ws_remaining = wb.create_sheet("Remaining Errors")
    ws_remaining.append(["Set Abbreviation", "Unit / Entry Point", "Error Type"])
    for row in remaining_errors:
        ws_remaining.append(list(row))
    wb.save(ERROR_FILE)
    print(f"\nSaved {len(remaining_errors)} remaining error(s) to '{ERROR_FILE}' > 'Remaining Errors' sheet.")

print("\nDone.")
